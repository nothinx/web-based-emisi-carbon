"""Router laporan: export Excel (.xlsx) sebuah CalculationRun + methodology appendix.

Generik lintas-domain: dibangun dari `CalculationResult` immutable (snapshot beku),
jadi laporan deterministik & dapat dipertahankan secara akademik. PDF tidak dibuat
server-side (lihat pyproject: laporan HTML cetak di frontend → browser Print→PDF,
agar tanpa native lib & parity dengan demo statis).
"""

from __future__ import annotations

import io
import uuid

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.api.deps import CurrentUser, SessionDep
from app.domains.base import build_methodology
from app.models.project import CalculationRun, Project
from app.models.registry import GWPSet

router = APIRouter(prefix="/reports", tags=["reports"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _load_run(session, run_id: uuid.UUID):
    run = (
        await session.execute(
            select(CalculationRun)
            .where(CalculationRun.id == run_id)
            .options(selectinload(CalculationRun.results))
        )
    ).scalar_one_or_none()
    if run is None:
        raise HTTPException(status_code=404, detail="Run tidak ditemukan")
    project = (
        await session.execute(select(Project).where(Project.id == run.project_id))
    ).scalar_one()
    gwp = (
        await session.execute(select(GWPSet).where(GWPSet.id == run.gwp_set_id))
    ).scalar_one()
    return run, project, gwp


def _build_workbook(run, project, gwp):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)
    results = list(run.results)
    total = sum(r.co2e_kg for r in results)

    wb = Workbook()

    # --- Sheet 1: Ringkasan ---
    ws = wb.active
    ws.title = "Ringkasan"
    rows = [
        ("Laporan Emisi — Carbon Engine", ""),
        ("Proyek", project.name),
        ("Domain", project.domain),
        ("Region", project.region),
        ("Base year", project.base_year or "—"),
        ("Periode", f"{project.reporting_period_start or '—'} → {project.reporting_period_end or '—'}"),
        ("GWP set", f"{gwp.name} (GWP-{gwp.horizon_years})"),
        ("Run ID", str(run.id)),
        ("Metode ketidakpastian", run.uncertainty_method),
        ("Total (kg CO2e)", round(total, 4)),
        ("Total (t CO2e)", round(total / 1000, 6)),
    ]
    for r0, r1 in rows:
        ws.append([r0, r1])
    ws["A1"].font = Font(bold=True, size=13)
    for i in range(2, len(rows) + 1):
        ws[f"A{i}"].font = bold

    # Rollup per scope (bila ada scope di snapshot).
    by_scope: dict[int, float] = {}
    for r in results:
        sc = r.factor_snapshot.get("category", {}).get("scope")
        if sc is not None:
            by_scope[sc] = by_scope.get(sc, 0.0) + r.co2e_kg
    if by_scope:
        ws.append([])
        ws.append(["Rollup per Scope (GHG Protocol)", "kg CO2e", "t CO2e", "share %"])
        ws[f"A{ws.max_row}"].font = bold
        for sc in sorted(by_scope):
            v = by_scope[sc]
            ws.append([f"Scope {sc}", round(v, 4), round(v / 1000, 6),
                       round(v / total * 100, 2) if total else 0])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    # --- Sheet 2: Hasil per aktivitas (immutable) ---
    wsr = wb.create_sheet("Hasil")
    headers = ["Kategori", "Scope", "Gas", "Strategy", "co2e_kg",
               "sd", "ci_low", "ci_high", "Faktor", "Unit", "Region", "Versi"]
    wsr.append(headers)
    for c in range(1, len(headers) + 1):
        wsr.cell(row=1, column=c).font = bold
    for r in results:
        s = r.factor_snapshot
        cat = s.get("category", {})
        unc = r.co2e_uncertainty or {}
        wsr.append([
            cat.get("name"), cat.get("scope"), s.get("gas", {}).get("symbol"),
            r.strategy_used, round(r.co2e_kg, 6),
            unc.get("sd"), unc.get("ci_low"), unc.get("ci_high"),
            s.get("value"), s.get("unit"), s.get("region"), s.get("version"),
        ])
    for col, w in {"A": 26, "D": 18, "E": 14, "I": 12, "J": 14, "K": 14}.items():
        wsr.column_dimensions[col].width = w

    # --- Sheet 3: Methodology appendix (faktor unik + sitasi) ---
    wsm = wb.create_sheet("Methodology")
    mheaders = ["Kategori", "Scope", "Gas", "Nilai faktor", "Unit", "Versi", "Region",
                "GWP diterapkan", "Tier", "Sumber", "Publisher", "Tahun", "Tier sumber",
                "URL", "Ketidakpastian"]
    wsm.append(mheaders)
    for c in range(1, len(mheaders) + 1):
        wsm.cell(row=1, column=c).font = bold
    for m in build_methodology(results):
        src = m.get("source", {})
        unc = m.get("uncertainty") or {}
        unc_txt = (
            f"±{unc['uncertainty_pct']}%" if unc.get("uncertainty_pct") is not None
            else (f"{unc.get('dist_type')} {unc.get('dist_params')}" if unc.get("dist_type") else "—")
        )
        wsm.append([
            m.get("category_name"), m.get("scope"), m.get("gas"), m.get("value"),
            m.get("unit"), m.get("version"), m.get("region"), m.get("gwp_applied"),
            m.get("tier"), src.get("name"), src.get("publisher"), src.get("year"),
            src.get("credibility_tier"), src.get("url"), unc_txt,
        ])
    for col, w in {"A": 26, "J": 34, "K": 24, "N": 40, "O": 22}.items():
        wsm.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/{run_id}.xlsx")
async def report_xlsx(run_id: uuid.UUID, session: SessionDep, _: CurrentUser):
    run, project, gwp = await _load_run(session, run_id)
    buf = _build_workbook(run, project, gwp)
    fname = f"laporan-{project.domain}-{str(run.id)[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
