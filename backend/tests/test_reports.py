"""Test export laporan Excel: endpoint .xlsx + isi methodology appendix."""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook


async def _auth(client) -> dict:
    await client.post("/auth/register", json={"email": "rep@uji.id", "password": "rahasia123"})
    tok = await client.post("/auth/token", data={"username": "rep@uji.id", "password": "rahasia123"})
    return {"Authorization": f"Bearer {tok.json()['access_token']}"}


async def _make_org_run(client, headers) -> str:
    r = await client.post(
        "/domains/organizational/calculate",
        headers=headers,
        json={
            "gwp_set_name": "AR6",
            "base_year": 2024,
            "inputs": {
                "facilities": [
                    {"name": "Pabrik A", "region": "ID-Jamali", "electricity_kwh": 100000},
                    {"name": "Pabrik B", "region": "ID-Sumatera", "diesel_stationary_l": 1000},
                ],
                "business_travel_air_pkm": 50000,
            },
        },
    )
    assert r.status_code == 200
    return r.json()["run_id"]


@pytest.mark.asyncio
async def test_report_xlsx_download(client):
    headers = await _auth(client)
    run_id = await _make_org_run(client, headers)

    r = await client.get(f"/reports/{run_id}.xlsx", headers=headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Ringkasan", "Hasil", "Methodology"]

    # Methodology appendix memuat sitasi sumber (mis. DEFRA/PLN) — bukan kosong.
    meth = wb["Methodology"]
    names = [meth.cell(row=row, column=10).value for row in range(2, meth.max_row + 1)]
    assert any(n and ("DEFRA" in n or "PLN" in n) for n in names)
    # Ada baris faktor grid regional (ID-Jamali) di appendix.
    regions = [meth.cell(row=row, column=7).value for row in range(2, meth.max_row + 1)]
    assert "ID-Jamali" in regions


@pytest.mark.asyncio
async def test_report_xlsx_requires_auth(client):
    headers = await _auth(client)
    run_id = await _make_org_run(client, headers)
    r = await client.get(f"/reports/{run_id}.xlsx")  # tanpa token
    assert r.status_code == 401


@pytest.mark.asyncio
async def test_report_xlsx_unknown_run_404(client):
    headers = await _auth(client)
    r = await client.get(
        "/reports/00000000-0000-0000-0000-000000000000.xlsx", headers=headers
    )
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_report_methodology_in_calc_response(client):
    """Report payload domain calculate membawa methodology (untuk laporan HTML/PDF)."""
    headers = await _auth(client)
    r = await client.post(
        "/domains/personal/calculate",
        headers=headers,
        json={"region": "ID-Jamali", "gwp_set_name": "AR6",
              "inputs": {"electricity_kwh_month": 250, "beef_kg_week": 1}},
    )
    report = r.json()["report"]
    assert report["methodology"]
    codes = {m["category"] for m in report["methodology"]}
    assert {"elec_grid", "food_beef"} <= codes
    # tiap entri membawa sitasi sumber.
    assert all("source" in m and m["source"].get("name") for m in report["methodology"])
